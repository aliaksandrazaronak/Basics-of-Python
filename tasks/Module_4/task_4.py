from openpyxl import load_workbook, Workbook

wb = load_workbook("task4_worksheet.xlsx")
ws = wb.active

new_wb = Workbook()
new_ws = new_wb.active

for col_index, column in enumerate(ws.columns, start=1):
    for row_index, row in enumerate(ws.rows, start=1):
        new_ws.cell(col_index, row_index).value = ws.cell(row_index, col_index).value
new_wb.save("task4_new_worksheet.xlsx")
